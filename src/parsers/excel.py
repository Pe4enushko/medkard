"""
excel.py — append audit results to an xlsx workbook.

Column layout (left to right):
  A  Специализация   — Врач.SPECIALIZATION
  B  Дата приема     — Прием.DATE
  C  Данные карты    — Пациент + Врач + Прием dicts
  D  Данные осмотра  — ДанныеОсмотра list
  E  Услуги          — Услуги list
  F  Диагнозы        — Диагнозы list
  G  formal_structure
  H  diagnosis

Legacy column layout (left to right):
  A  Данные карты    — full visit dump
  B  formal_structure
  C  diagnosis

Usage::
    from parsers.excel import AuditExcelWriter
    from audit.models import DiagnosisAuditResult, FormalStructureResult

    writer = AuditExcelWriter("results.xlsx")
    writer.append(
        visit=raw_visit_dict,
        formal=FormalStructureResult(...),
        diagnosis=DiagnosisAuditResult(...),
    )
"""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
import io
import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from audit.models import FormalStructureResult
from parsers.inspection_order import reorder_inspection_data

_HEADERS = [
    "Специализация",
    "Дата приема",
    "Данные карты",
    "Данные осмотра",
    "Услуги",
    "Диагнозы",
    "Проверка по приказам МЗ",
    "Проверка по клин.рекоммендациям",
    "Проверка кодирования МКБ",
]
_COLUMN_WIDTHS = {
    "A": 25,
    "B": 15,
    "C": 60,
    "D": 178,
    "E": 60,
    "F": 60,
    "G": 80,
    "H": 100,
    "I": 100,
}
_AUTOFILTER_RANGE = "A1:B1"

_LEGACY_HEADERS = ["visits", "formal", "diagnosis"]
_LEGACY_COLUMN_WIDTHS = {"A": 100, "B": 80, "C": 100}
_LEGACY_AUTOFILTER_RANGE = "A1:A1"

logger = logging.getLogger(__name__)


def _pretty(obj: Any) -> str:
    if hasattr(obj, "pretty_format") and callable(obj.pretty_format):
        return obj.pretty_format()

    if isinstance(obj, list) and all(
        hasattr(item, "pretty_format") and callable(item.pretty_format)
        for item in obj
    ):
        if not obj:
            return "Нет"
        return "\n\n".join(
            f"{idx}.\n{item.pretty_format()}"
            for idx, item in enumerate(obj, start=1)
        )

    if is_dataclass(obj):
        obj = asdict(obj)

    return _format_value(obj)


def _format_value(value: Any, indent: int = 0) -> str:
    prefix = " " * indent

    if is_dataclass(value):
        value = asdict(value)

    if isinstance(value, dict):
        if not value:
            return f"{prefix}—"

        lines: list[str] = []
        for key, item in value.items():
            label = str(key)
            if _is_scalar(item):
                lines.append(f"{prefix}{label}: {_format_scalar(item)}")
            else:
                lines.append(f"{prefix}{label}:")
                lines.append(_format_value(item, indent + 2))
        return "\n".join(lines)

    if isinstance(value, list):
        if not value:
            return f"{prefix}—"

        lines: list[str] = []
        for idx, item in enumerate(value, start=1):
            if _is_scalar(item):
                lines.append(f"{prefix}{idx}. {_format_scalar(item)}")
            else:
                lines.append(f"{prefix}{idx}.")
                lines.append(_format_value(item, indent + 2))
        return "\n".join(lines)

    return f"{prefix}{_format_scalar(value)}"


def _is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _format_scalar(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "да" if value else "нет"
    return str(value)


def _card_data_text(visit: dict[str, Any]) -> str:
    """Render Пациент + Врач + Прием as a single text block."""
    parts: list[str] = []
    for key in ("Пациент", "Врач", "Прием"):
        val = visit.get(key)
        if val is not None:
            parts.append(f"{key}:\n{_format_value(val, indent=2)}")
    return "\n\n".join(parts) if parts else "—"


def _build_row(
    visit: dict[str, Any],
    formal: FormalStructureResult,
    diagnosis: Any,
    icd_check: Any = None,
    *,
    legacy: bool = False,
    order_tokens: list[str] | None = None,
) -> list[str]:
    if legacy:
        return [_pretty(visit), _pretty(formal), _pretty(diagnosis)]

    specialization = (visit.get("Врач") or {}).get("SPECIALIZATION") or "—"
    visit_date = (visit.get("Прием") or {}).get("DATE") or "—"
    inspection = visit.get("ДанныеОсмотра") or []
    if order_tokens:
        inspection = reorder_inspection_data(inspection, order_tokens)
    return [
        specialization,
        visit_date,
        _card_data_text(visit),
        _pretty(inspection),
        _pretty(visit.get("Услуги") or []),
        _pretty(visit.get("Диагнозы") or []),
        _pretty(formal),
        _pretty(diagnosis),
        _pretty(icd_check or []),
    ]


def build_workbook_bytes(
    rows: list[tuple[dict[str, Any], FormalStructureResult, Any, Any]],
    *,
    legacy: bool = False,
    order_tokens: list[str] | None = None,
) -> bytes:
    """Build an xlsx workbook in memory from (visit, formal, diagnosis, icd_check)
    tuples and return its bytes — no disk I/O, for use in per-request API responses.
    """
    wb = Workbook()
    ws = wb.active
    headers = _LEGACY_HEADERS if legacy else _HEADERS
    widths = _LEGACY_COLUMN_WIDTHS if legacy else _COLUMN_WIDTHS
    autofilter = _LEGACY_AUTOFILTER_RANGE if legacy else _AUTOFILTER_RANGE

    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.auto_filter.ref = autofilter

    for visit, formal, diagnosis, icd_check in rows:
        row = _build_row(visit, formal, diagnosis, icd_check, legacy=legacy, order_tokens=order_tokens)
        ws.append(row)
        new_row = ws.max_row
        for col_idx in range(1, len(row) + 1):
            ws.cell(row=new_row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class AuditExcelWriter:
    """Append audit results to an xlsx file, creating it with a header row if absent.

    Args:
        path:         Path to the xlsx output file (created automatically if missing).
        legacy:       Use the legacy 3-column layout (visits / formal / diagnosis).
        order_tokens: Optional flat token list used to reorder each visit's
                       ДанныеОсмотра before rendering (see reorder_inspection_data).
    """

    def __init__(
        self,
        path: str | Path,
        *,
        legacy: bool = False,
        order_tokens: list[str] | None = None,
    ) -> None:
        self._path = Path(path)
        self._legacy = legacy
        self._order_tokens = order_tokens

    def _open_or_create(self) -> tuple[Workbook, Worksheet]:
        if self._path.exists():
            wb = openpyxl.load_workbook(self._path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
        if self._legacy:
            headers = _LEGACY_HEADERS
            widths = _LEGACY_COLUMN_WIDTHS
            autofilter = _LEGACY_AUTOFILTER_RANGE
        else:
            headers = _HEADERS
            widths = _COLUMN_WIDTHS
            autofilter = _AUTOFILTER_RANGE
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        ws.auto_filter.ref = autofilter
        return wb, ws  # type: ignore[return-value]

    def append(
        self,
        visit: dict[str, Any],
        formal: FormalStructureResult,
        diagnosis: Any,
        icd_check: Any = None,
    ) -> None:
        """Append one result row and save the workbook.

        Args:
            visit:     Raw visit dict (source JSON from 1C).
            formal:    Formal-structure audit result.
            diagnosis: Diagnosis audit result(s) for this visit.
        """
        try:
            row = _build_row(
                visit, formal, diagnosis, icd_check,
                legacy=self._legacy, order_tokens=self._order_tokens,
            )
            wb, ws = self._open_or_create()
            ws.append(row)
            new_row = ws.max_row
            for col_idx in range(1, len(row) + 1):
                ws.cell(row=new_row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[new_row].height = None
            wb.save(self._path)
            logger.info("📊 EXCEL APPEND OK path=%s", self._path)
        except Exception:
            logger.exception("📊 EXCEL APPEND FAILED path=%s", self._path)
            raise

    def rows_count(self) -> int:
        """Return current number of rows in the active worksheet (0 if file absent)."""
        if not self._path.exists():
            return 0
        wb = openpyxl.load_workbook(self._path)
        try:
            return wb.active.max_row
        finally:
            wb.close()
