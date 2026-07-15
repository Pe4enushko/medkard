"""
audit/excel_formatter.py — export done_cards rows to an xlsx workbook.

Reads every row from the done_cards table (or a specific subset by guid)
and appends them to an Excel file via AuditExcelWriter.  The pipeline no
longer writes Excel directly; call this after the pipeline completes (or
any time) to regenerate / update the workbook from the DB.

Usage::
    from audit.excel_formatter import ExcelFormatter

    async with ExcelFormatter("audit_results.xlsx") as fmt:
        written = await fmt.export_all()
        print(f"wrote {written} rows")
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from parsers.excel import AuditExcelWriter
from reporting.result_parser import build_manifest_meta as _build_manifest_meta
from reporting.result_parser import parse_diagnosis as _parse_diagnosis
from reporting.result_parser import parse_formal as _parse_formal
from reporting.result_parser import parse_icd_check as _parse_icd_check
from storage.base import BaseStorage
from storage.guidelines_storage import GuidelinesStorage

logger = logging.getLogger(__name__)


class _DoneCardsReader(BaseStorage):
    @staticmethod
    def _decode_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        for row in rows:
            if isinstance(row["card_data"], str):
                row["card_data"] = json.loads(row["card_data"])
        return rows

    async def fetch_all(self) -> list[dict[str, Any]]:
        async with self._pool.connection() as conn:
            cur = await conn.execute(
                "SELECT id, card_guid, card_data::text, formal_result, diag_result, icd_check_result "
                "FROM done_cards ORDER BY id"
            )
            return self._decode_rows(await cur.fetchall())

    async def fetch_by_guids(self, guids: set[str]) -> list[dict[str, Any]]:
        async with self._pool.connection() as conn:
            cur = await conn.execute(
                "SELECT id, card_guid, card_data::text, formal_result, diag_result, icd_check_result "
                "FROM done_cards WHERE card_guid = ANY(%(guids)s) ORDER BY id",
                {"guids": list(guids)},
            )
            return self._decode_rows(await cur.fetchall())

    async def fetch_by_period(
        self, date_from: datetime, date_to: datetime, organization_id: str
    ) -> list[dict[str, Any]]:
        if not organization_id:
            raise ValueError("organization_id is required: a report must be scoped to one org")
        async with self._pool.connection() as conn:
            cur = await conn.execute(
                "WITH cards AS ("
                "  SELECT id, card_guid, card_data, formal_result, diag_result, icd_check_result, "
                "         to_date(card_data -> 'Прием' ->> 'DATE', 'DD.MM.YYYY') AS visit_date "
                "  FROM done_cards "
                "  WHERE ignored = FALSE "
                "    AND broken = FALSE "
                "    AND organization_id = %(org_id)s::uuid"
                ") "
                "SELECT id, card_guid, card_data::text, formal_result, diag_result, icd_check_result "
                "FROM cards "
                "WHERE visit_date >= %(from)s::date AND visit_date < %(to)s::date "
                "ORDER BY visit_date, id",
                {"from": date_from, "to": date_to, "org_id": organization_id},
            )
            return self._decode_rows(await cur.fetchall())


def _existing_guids_in_excel(excel: AuditExcelWriter) -> set[str]:
    """Return the set of appointment GUIDs already present in the Excel sheet.

    The GUID appears as «GUID: <value>» inside the «Данные карты» column.
    """
    import openpyxl

    path = excel._path
    if not path.exists():
        return set()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        existing: set[str] = set()
        card_col = None
        for cell in ws[1]:
            if cell.value in ("Данные карты", "input"):
                card_col = cell.column
                break
        if card_col is None:
            return set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_value = row[card_col - 1]
            if not cell_value:
                continue
            for line in str(cell_value).splitlines():
                stripped = line.strip()
                if stripped.startswith("GUID:"):
                    guid = stripped[len("GUID:"):].strip().lower()
                    if guid:
                        existing.add(guid)
                    break
        return existing
    finally:
        wb.close()


class ExcelFormatter:
    """Async context-manager that exports done_cards rows to an xlsx file.

    Args:
        excel_path:    Path to the output xlsx file (created if absent).
        legacy:        Use the legacy 3-column layout (visits / formal / diagnosis).
        order_tokens:  Optional canonical field order for ДанныеОсмотра; ``None``
                       disables reordering.
    """

    def __init__(
        self,
        excel_path: str | Path,
        *,
        legacy: bool = False,
        order_tokens: list[str] | None = None,
    ) -> None:
        self._excel = AuditExcelWriter(excel_path, legacy=legacy, order_tokens=order_tokens)
        self._reader = _DoneCardsReader()

    async def __aenter__(self) -> "ExcelFormatter":
        await self._reader.__aenter__()
        return self

    async def __aexit__(self, *args: object) -> None:
        await self._reader.__aexit__(*args)

    async def export_all(self) -> int:
        """Write every done_cards row to Excel. Returns number of rows written."""
        rows = await self._reader.fetch_all()
        return self._write_rows(rows, await self._load_meta())

    async def export_period(
        self, date_from: datetime, date_to: datetime, organization_id: str
    ) -> int:
        """Write *organization_id*'s rows whose appointment date is in [date_from, date_to).

        ``organization_id`` is mandatory: a report file is always scoped to a single
        organization so cards from different orgs can never land in the same file.
        """
        rows = await self._reader.fetch_by_period(date_from, date_to, organization_id)
        return self._write_rows(rows, await self._load_meta())

    async def export_guids(self, guids: set[str]) -> int:
        """Write only the rows matching *guids*. Returns number of rows written."""
        rows = await self._reader.fetch_by_guids(guids)
        return self._write_rows(rows, await self._load_meta())

    async def _load_meta(self) -> dict:
        async with GuidelinesStorage() as store:
            return _build_manifest_meta(await store.all())

    def _write_rows(self, rows: list[dict[str, Any]], manifest_meta: dict) -> int:
        existing = _existing_guids_in_excel(self._excel)
        written = 0
        for row in rows:
            guid = (row["card_guid"] or "").lower()
            if guid and guid in existing:
                logger.debug("📊 skipping already exported card guid=%s", guid)
                continue
            visit = row["card_data"]
            formal = _parse_formal(row["formal_result"])
            diagnosis = _parse_diagnosis(row["diag_result"], manifest_meta)
            icd_check = _parse_icd_check(row.get("icd_check_result") or [])
            self._excel.append(visit=visit, formal=formal, diagnosis=diagnosis, icd_check=icd_check)
            logger.debug("📊 exported card id=%s guid=%s", row["id"], guid)
            written += 1
        logger.info("📊 ExcelFormatter exported %d row(s)", written)
        return written
