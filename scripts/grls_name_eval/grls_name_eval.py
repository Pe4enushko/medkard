#!/usr/bin/env python
"""Эвал способов поиска по реестру лекарств — на настоящем Postgres.

Вопрос: нужен ли вектор поверх лексики, и какая лексика лучше. От ответа зависит,
есть ли в drug_registry движка колонка embedding, HNSW на 39 тыс. строк и фаза
эмбеддинга в синке (спека engine
docs/superpowers/specs/2026-08-20-grls-integration-design.md, §4.5 — другой
репозиторий, ветка grls-integration).

Считается НЕ приближением, а тем же движком, что будет в проде: скрипт создаёт
ВРЕМЕННУЮ таблицу в БД из .env, кладёт туда корпус, строит текстовые индексы
(GIN pg_trgm, GIN tsvector) и ранжирует запросы средствами Postgres —
`similarity()`, `ts_rank()`, `websearch_to_tsquery('russian', …)`, `pgvector`.
Временная таблица живёт внутри сессии и исчезает при выходе; в БД ничего не
остаётся.

СРАВНИВАЕМЫЕ МЕТОДЫ:
  trigram      — pg_trgm по всему индексируемому тексту;
  tsv          — to_tsvector('russian') по «остальному» (ФТГ, формы, отпуск,
                 держатель), без названий;
  trgm+tsv     — ПРОДАКШН-КАНДИДАТ: триграммы по названиям (торговое + МНН) плюс
                 tsv по остальному, слияние через RRF. То, что планируется в Искре;
  vector       — только косинус по эмбеддингу (pgvector);
  hybrid       — vector + trigram с весом, как в миграции 084: r_vec + 0.3*r_trgm;
  hybrid_even  — vector + trigram без весов: вес 0.3 подобран под клинреки и на
                 лекарствах может не подойти;
  trgm+tsv+vec — продакшн-кандидат плюс вектор: показывает, что вектор ДОБАВЛЯЕТ
                 к тому, что и так будет.

ЧТО ИНДЕКСИРУЕТСЯ (--index):
  blob (по умолчанию) — search_blob: торговое | МНН | ФТГ | лекформы | отпуск |
                        держатель. Так делает спека §4.5;
  name               — только торговое наименование (узкий эксперимент).
Режим меняет выводы: с blob'ом МНН и ФТГ лежат ВНУТРИ индексируемого текста и
находятся лексикой, поэтому у вектора остаётся меньше работы.

КЛАССЫ ЗАПРОСОВ (см. build_queries):
  L0 exact       — точное название: проверка вменяемости, ожидается ~100%;
  L1 typo        — регистр, ё→е, снятие ®, лишние пробелы;
  L2 typo1       — одна опечатка в названии;
  L3 typo3       — три опечатки в названии;
  L4 phonetic    — типичные русские искажения на слух (о↔а, е↔и, тс↔ц…);
  L5 inn2trade   — по МНН найти торговое (семантическая связь);
  L6 ftg exact   — точная ФТГ;
  L7 ftg typo    — ФТГ с опечатками;
  L8 ftg partial — фрагмент ФТГ («противоопухолевое», «средство растительного»).

МНОЖЕСТВЕННЫЙ ОТВЕТ. У классов L6–L8 правильных ответов много: одну ФТГ делят
десятки препаратов. Поэтому эталон — множество, а ранг считается по ПЕРВОМУ
релевантному. Для L0–L5 множество из одного элемента, метрика читается
одинаково: «хотя бы один релевантный попал в топ-k». В отчёте печатается
средний размер множества (|gold|): классы с разным |gold| между собой
несравнимы, сравнивать надо МЕТОДЫ внутри класса.

Запуск (нужны доступ к эндпоинту эмбеддингов и Postgres из .env):

    python grls_name_eval.py --from-zip ~/projects/grls2026-08-17-1.zip
    python grls_name_eval.py --from-db "postgresql://user@host/medkard"

Эндпоинт, ключ, модель и параметры Postgres берутся из .env репозитория
(EMBEDDING_BASE_URL или OPENAI_BASE_URL, OPENAI_API_KEY, EMBEDDING_MODEL,
POSTGRES_*) — как в medkard; перебиваются флагами.

ВАЖНО: выгрузка открытых данных (data-*.csv / data-*.json) для эвала НЕ
годится — в ней ноль торговых наименований и МНН (та же спека, §3.3). Имена —
только xlsx-архив ГРЛС или уже загруженная таблица grls_registry.
"""
from __future__ import annotations

import argparse
import json
import os
import random
import re
import sys
import unicodedata
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path


# .env репозитория — грузим ДО чтения переменных, чтобы дефолты в argparse
# уже видели значения. Ищем вверх от скрипта: scripts/grls_name_eval/ → корень.
def _load_env() -> Path | None:
    for parent in Path(__file__).resolve().parents:
        candidate = parent / ".env"
        if candidate.is_file():
            try:
                from dotenv import load_dotenv
                load_dotenv(candidate)
            except ImportError:  # без python-dotenv разбираем сами: KEY=VALUE
                for line in candidate.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, value = line.partition("=")
                    os.environ.setdefault(key.strip(), value.strip().strip("\"'"))
            return candidate
    return None


ENV_PATH = _load_env()

# Переменные — как у medkard (src/RAG/retrieval/embeddings.py): база берётся из
# EMBEDDING_BASE_URL, иначе OPENAI_BASE_URL; ключ — OPENAI_API_KEY.
DEFAULT_BASE_URL = (
    os.getenv("EMBEDDING_BASE_URL")
    or os.getenv("OPENAI_BASE_URL")
    or "https://api.openai.com/v1"
)
DEFAULT_API_KEY = os.getenv("OPENAI_API_KEY", "")
DEFAULT_MODEL = os.getenv("EMBEDDING_MODEL", "Qwen/Qwen3-Embedding-0.6B")

METHODS = ("trigram", "tsv", "trgm+tsv", "vector", "hybrid", "hybrid_even", "trgm+tsv+vec")

# Вес слияния — как в 084_clinical_guideline_name_embedding.sql
W_TSV = 0.3
# RRF: score = Σ 1/(K + rank). Так сливает ветки medkard (BM25 + вектор, см.
# CLAUDE.md). Сумма рангов для этого не годится: если ветка по запросу пустая
# (ts_rank = 0 у всех, запрос — название), её ранг произволен и подмешивает в
# сумму чистый шум, утягивая вниз даже точные попадания другой ветки. У RRF
# вклад плохого ранга затухает, и пустая ветка почти ничего не портит.
RRF_K = 60
EMBED_BATCH = 64
TOP_K = 10
SEED = 20260820
# Порог для L5: пары МНН↔торговое с триграммным сходством выше — лексические
# («Церебролизин» ← «церебролизин»), для проверки семантики бесполезны.
INN_LEXICAL_MAX = 0.30
# ФТГ короче — обрывки вроде «~» или «прочие»; в классы L6–L8 не берём.
FTG_MIN_CHARS = 12

# ─────────────────────────── нормализация ───────────────────────────
# Зеркало grls_norm() канона (спека medkard §3.1): lower, схлопывание пробелов,
# удаление ®™© и кавычек, ё→е, ~→пусто. Держать идентичной канону.
_JUNK = re.compile(r"[®™©\"«»„“”'`]")
_SPACES = re.compile(r"\s+")


def grls_norm(text: str | None) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", text)
    text = _JUNK.sub("", text)
    text = text.replace("ё", "е").replace("Ё", "Е").replace("~", "")
    return _SPACES.sub(" ", text).strip().lower()


# Триграммы нужны только для отбора пар в L5 (лексическая похожесть МНН и
# торгового). Ранжирование считает Postgres — приближение в нём не участвует.
_WORD_SPLIT = re.compile(r"[^0-9a-zа-я]+")


def _trigrams(text: str) -> set[str]:
    out: set[str] = set()
    for word in _WORD_SPLIT.split(grls_norm(text)):
        if not word:
            continue
        padded = f"  {word} "
        for i in range(len(padded) - 2):
            out.add(padded[i:i + 3])
    return out


def trgm_similarity(a: str, b_trg: set[str]) -> float:
    a_trg = _trigrams(a)
    if not a_trg or not b_trg:
        return 0.0
    inter = len(a_trg & b_trg)
    return inter / len(a_trg | b_trg) if inter else 0.0


# ─────────────────────────── производные формы ───────────────────────────
# Те же правила, что в medkard src/grls/normalize.py: первый сегмент до запятой
# = лекарственная форма, последний сегмент после ' - ' = условия отпуска.
def split_forms(forms_raw: str | None) -> list[str]:
    return [p.strip() for p in (forms_raw or "").split(";") if p.strip()]


def _unique(items: list[str]) -> list[str]:
    seen, out = set(), []
    for it in items:
        if it and it not in seen:
            seen.add(it)
            out.append(it)
    return out


def derive_dosage_forms(forms: list[str]) -> list[str]:
    return _unique([el.split(",", 1)[0].strip() for el in forms if not el.startswith("-")])


def derive_dispensing(forms: list[str]) -> list[str]:
    return _unique([el.rsplit(" - ", 1)[1].strip() for el in forms if " - " in el])


def names_text(rec: dict) -> str:
    """Названия: торговое + МНН. По ним в проде идут триграммы."""
    return " | ".join(p.strip() for p in (rec.get("trade_name"), rec.get("inn_name"))
                      if p and p.strip() and p.strip() != "~")


def rest_text(rec: dict) -> str:
    """Всё остальное: ФТГ, лекформы, отпуск, держатель. По нему в проде идёт tsv."""
    forms = split_forms(rec.get("forms_raw"))
    parts = [
        rec.get("pharm_group"),
        ", ".join(derive_dosage_forms(forms)[:6]),
        ", ".join(derive_dispensing(forms)[:4]),
        rec.get("holder"),
    ]
    return " | ".join(p.strip() for p in parts if p and p.strip() and p.strip() != "~")


def search_blob(rec: dict) -> str:
    """Поисковый текст записи целиком — как search_blob() в спеке engine §4.5.

    forms_raw целиком НЕ идёт: у препарата с десятком упаковок это килобайты,
    размывающие и лексический, и векторный сигнал; суть уже в производных.
    """
    return " | ".join(p for p in (names_text(rec), rest_text(rec)) if p)


# ─────────────────────────── чтение корпуса ───────────────────────────
# Раскладка листа — как в парсере medkard: строка 5 заголовки, данные с 7-й,
# колонки C..Q. Внутри среза: 4 держатель, 6 торговое, 7 МНН, 8 формы, 11 ФТГ.
_COL_HOLDER, _COL_TRADE, _COL_INN, _COL_FORMS, _COL_FTG = 4, 6, 7, 8, 11


def load_from_zip(path: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("нужен openpyxl: pip install openpyxl")

    first, n_cols = 2, 15
    rows: list[dict] = []
    with zipfile.ZipFile(path) as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not names:
            sys.exit(f"в архиве {path} нет xlsx")
        tmp = Path("/tmp/grls_eval_sheets")
        tmp.mkdir(exist_ok=True)
        for name in names:
            target = tmp / Path(name).name
            with zf.open(name) as src, open(target, "wb") as dst:
                dst.write(src.read())
            wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i < 7:
                    continue
                cells = (tuple(row or ()) + (None,) * (first + n_cols))[first:first + n_cols]

                def cell(idx: int) -> str | None:
                    v = cells[idx]
                    if v is None:
                        return None
                    v = str(v).strip()
                    return v if v and v != "~" else None

                trade = cell(_COL_TRADE)
                if not trade:
                    continue
                rows.append({
                    "trade_name": trade,
                    "inn_name": cell(_COL_INN),
                    "pharm_group": cell(_COL_FTG),
                    "forms_raw": cell(_COL_FORMS),
                    "holder": cell(_COL_HOLDER),
                })
            wb.close()
            print(f"  прочитан {Path(name).name}: всего строк {len(rows)}", flush=True)
    return rows


def load_from_db(dsn: str) -> list[dict]:
    import psycopg
    with psycopg.connect(dsn) as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT trade_name, inn_name, pharm_group, forms_raw, holder "
            "FROM grls_registry WHERE NOT is_substance AND trade_name IS NOT NULL"
        )
        cols = [d.name for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def dedupe(rows: list[dict]) -> list[dict]:
    """Один препарат = одно уникальное нормализованное торговое наименование.

    В реестре у одного названия десятки строк (разные упаковки и редакции РУ);
    для эвала поиска они шум: цель — найти ПРЕПАРАТ, а не строку. Побочный
    эффект: омонимы разных производителей схлопываются в одну запись.
    """
    seen: dict[str, dict] = {}
    for r in rows:
        key = grls_norm(r["trade_name"])
        if not key:
            continue
        cur = seen.get(key)
        if cur is None:
            seen[key] = r
        elif (bool(r.get("inn_name")) + bool(r.get("pharm_group")) >
              bool(cur.get("inn_name")) + bool(cur.get("pharm_group"))):
            seen[key] = r  # предпочитаем более полную запись
    return list(seen.values())


# ─────────────────────────── зашумление ───────────────────────────
_RU = "абвгдежзийклмнопрстуфхцчшщыэюя"
_PHONETIC = [
    ("о", "а"), ("а", "о"), ("е", "и"), ("и", "е"), ("тс", "ц"), ("ц", "тс"),
    ("дт", "т"), ("сс", "с"), ("лл", "л"), ("нн", "н"), ("ф", "в"), ("в", "ф"),
]


def _typo(word: str, rng: random.Random) -> str:
    """Одна опечатка: замена, перестановка соседей, удаление или вставка."""
    if len(word) < 3:
        return word
    kind = rng.choice(("sub", "swap", "del", "ins"))
    i = rng.randrange(1, len(word) - 1)
    if kind == "sub":
        return word[:i] + rng.choice(_RU) + word[i + 1:]
    if kind == "swap":
        return word[:i] + word[i + 1] + word[i] + word[i + 2:]
    if kind == "del":
        return word[:i] + word[i + 1:]
    return word[:i] + rng.choice(_RU) + word[i:]


def _phonetic(word: str, rng: random.Random) -> str:
    pairs = list(_PHONETIC)
    rng.shuffle(pairs)
    for src, dst in pairs:
        if src in word:
            return word.replace(src, dst, 1)
    return _typo(word, rng)


def _partial(text: str, rng: random.Random) -> str:
    """Непрерывное окно из 1–3 слов — так врач называет группу не целиком."""
    words = grls_norm(text).split()
    if len(words) <= 2:
        return " ".join(words)
    size = rng.randint(1, min(3, len(words) - 1))
    start = rng.randrange(0, len(words) - size + 1)
    return " ".join(words[start:start + size])


def build_queries(corpus: list[dict], per_level: int, rng: random.Random) -> list[dict]:
    """По каждому классу — per_level запросов с известным множеством ответов.

    gold_ids — индексы корпуса, считающиеся правильными. Для названий это один
    элемент, для ФТГ — все препараты той же группы.
    """
    named = [(i, r) for i, r in enumerate(corpus) if len(grls_norm(r["trade_name"])) >= 5]
    with_inn = [
        (i, r) for i, r in named
        if r.get("inn_name")
        and trgm_similarity(r["inn_name"], _trigrams(r["trade_name"])) < INN_LEXICAL_MAX
    ]
    ftg_members: dict[str, list[int]] = defaultdict(list)
    for i, r in enumerate(corpus):
        key = grls_norm(r.get("pharm_group"))
        if len(key) >= FTG_MIN_CHARS:
            ftg_members[key].append(i)
    with_ftg = [(i, r) for i, r in named if grls_norm(r.get("pharm_group")) in ftg_members]

    queries: list[dict] = []

    def add(level: str, pool: list, make, gold_of=None) -> None:
        if not pool:
            print(f"  ВНИМАНИЕ: класс {level} пуст — пропущен", file=sys.stderr)
            return
        for idx, rec in rng.sample(pool, min(per_level, len(pool))):
            q = make(rec)
            if not q or not grls_norm(q):
                continue
            gold = gold_of(rec) if gold_of else frozenset({idx})
            if gold:
                queries.append({"level": level, "query": q, "gold_ids": gold})

    by_ftg = lambda r: frozenset(ftg_members[grls_norm(r["pharm_group"])])

    add("L0 exact", named, lambda r: r["trade_name"])
    add("L1 typo", named, lambda r: grls_norm(r["trade_name"]).upper().replace("е", "ё", 1))
    add("L2 typo1", named, lambda r: _typo(grls_norm(r["trade_name"]), rng))
    add("L3 typo3", named,
        lambda r: _typo(_typo(_typo(grls_norm(r["trade_name"]), rng), rng), rng))
    add("L4 phonetic", named, lambda r: _phonetic(grls_norm(r["trade_name"]), rng))
    add("L5 inn2trade", with_inn, lambda r: r["inn_name"])
    add("L6 ftg exact", with_ftg, lambda r: r["pharm_group"], by_ftg)
    add("L7 ftg typo", with_ftg,
        lambda r: _typo(_typo(grls_norm(r["pharm_group"]), rng), rng), by_ftg)
    add("L8 ftg partial", with_ftg, lambda r: _partial(r["pharm_group"], rng), by_ftg)
    return queries


# ─────────────────────────── эмбеддинги ───────────────────────────
def embed(texts: list[str], base_url: str, api_key: str, model: str) -> list[list[float]]:
    out: list[list[float]] = []
    url = base_url.rstrip("/") + "/embeddings"
    for start in range(0, len(texts), EMBED_BATCH):
        batch = [t or " " for t in texts[start:start + EMBED_BATCH]]
        body = json.dumps({"model": model, "input": batch}).encode()
        req = urllib.request.Request(
            url, data=body,
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {api_key}"},
        )
        with urllib.request.urlopen(req, timeout=180) as resp:
            data = json.load(resp)
        out.extend(item["embedding"] for item in sorted(data["data"], key=lambda d: d["index"]))
        print(f"    эмбеддинг {min(start + EMBED_BATCH, len(texts))}/{len(texts)}", flush=True)
    return out


def _pgvector(v: list[float]) -> str:
    return "[" + ",".join(f"{x:.6f}" for x in v) + "]"


# ─────────────────────────── Postgres ───────────────────────────
def pg_dsn_from_env() -> str:
    from urllib.parse import quote
    user = os.getenv("POSTGRES_USER", "")
    pwd = os.getenv("POSTGRES_PASSWORD", "")
    auth = f"{quote(user)}:{quote(pwd)}@" if user else ""
    return (f"postgresql://{auth}{os.getenv('POSTGRES_HOST', 'localhost')}:"
            f"{os.getenv('POSTGRES_PORT', '5432')}/{os.getenv('POSTGRES_DB', 'medkard')}")


# Ранжирование целиком средствами Postgres.
#
# Ранги ВЕТОК — dense_rank, как в проде (084_clinical_guideline_name_embedding
# .sql:76-77). Это не косметика: с row_number мёртвая ветка (ts_rank = 0 у всех,
# запрос-название) всё равно обязана выдать кому-то ранги 1, 2, 3… и раздаёт их
# по id. В RRF такой ранг весит 1/61 — ровно столько же, сколько точное
# попадание в живой ветке, и записи с малым id систематически всплывают поверх
# настоящего ответа. Подпись дефекта: recall@1 = 0.06 при recall@10 = 0.95.
# dense_rank схлопывает всю массу нулей в один ранг, он достаётся и эталону
# тоже, становится общим слагаемым и сокращается.
#
# ИТОГОВАЯ позиция (r_trgm..r_prod_vec наружу) — наоборот, row_number с добивкой
# по id: при массовых ничьих rank отдал бы всей группе лучшую позицию и завысил
# метрику, а row_number даёт ту произвольную, но реалистичную позицию, которую
# вернул бы обычный запрос с LIMIT.
_RANK_SQL = """
WITH scored AS (
  SELECT id,
         similarity(doc_norm,   %(q)s)                             AS s_trgm,
         similarity(names_norm, %(q)s)                             AS s_names,
         ts_rank(rest_tsv, websearch_to_tsquery('russian', %(q)s)) AS s_tsv,
         1 - (emb <=> %(qv)s::vector)                              AS s_vec
  FROM eval_docs
), ranked AS (
  SELECT id, s_trgm, s_tsv, s_vec,
    dense_rank() OVER (ORDER BY s_trgm  DESC) AS d_trgm,
    dense_rank() OVER (ORDER BY s_names DESC) AS d_names,
    dense_rank() OVER (ORDER BY s_tsv   DESC) AS d_tsv,
    dense_rank() OVER (ORDER BY s_vec   DESC) AS d_vec
  FROM scored
), fused AS (
  SELECT id,
    row_number() OVER (ORDER BY s_trgm DESC, id) AS r_trgm,
    row_number() OVER (ORDER BY s_tsv  DESC, id) AS r_tsv,
    row_number() OVER (ORDER BY s_vec  DESC, id) AS r_vec,
    row_number() OVER (ORDER BY (1.0/(%(k)s + d_names) + 1.0/(%(k)s + d_tsv))
                       DESC, id)                                   AS r_prod,
    row_number() OVER (ORDER BY (d_vec + %(w)s * d_trgm) ASC, id)  AS r_hybrid,
    row_number() OVER (ORDER BY (d_vec + d_trgm) ASC, id)          AS r_even,
    row_number() OVER (ORDER BY (1.0/(%(k)s + d_names) + 1.0/(%(k)s + d_tsv)
                                 + 1.0/(%(k)s + d_vec)) DESC, id)  AS r_prod_vec
  FROM ranked
)
SELECT min(r_trgm), min(r_tsv), min(r_prod), min(r_vec),
       min(r_hybrid), min(r_even), min(r_prod_vec)
FROM fused WHERE id = ANY(%(gold)s)
"""


def build_table(conn, corpus, docs, doc_vecs, dim) -> None:
    with conn.cursor() as cur:
        for ext in ("pg_trgm", "vector"):
            try:
                cur.execute(f"CREATE EXTENSION IF NOT EXISTS {ext}")
                conn.commit()
            except Exception as e:
                conn.rollback()
                sys.exit(f"нет расширения {ext} и его не создать: {e}\n"
                         "эвал считает лексику и вектор средствами Postgres — без них никак")
        cur.execute(f"""
            CREATE TEMP TABLE eval_docs (
                id         int PRIMARY KEY,
                doc_norm   text,
                names_norm text,
                rest       text,
                rest_tsv   tsvector GENERATED ALWAYS AS
                           (to_tsvector('russian', coalesce(rest, ''))) STORED,
                emb        vector({dim})
            ) ON COMMIT PRESERVE ROWS
        """)
        cur.executemany(
            "INSERT INTO eval_docs (id, doc_norm, names_norm, rest, emb) "
            "VALUES (%s, %s, %s, %s, %s)",
            [(i, grls_norm(docs[i]), grls_norm(names_text(r)), rest_text(r),
              _pgvector(doc_vecs[i])) for i, r in enumerate(corpus)],
        )
        # Текстовые индексы — как будут в проде (спека §4.5).
        cur.execute("CREATE INDEX ON eval_docs USING GIN (doc_norm gin_trgm_ops)")
        cur.execute("CREATE INDEX ON eval_docs USING GIN (names_norm gin_trgm_ops)")
        cur.execute("CREATE INDEX ON eval_docs USING GIN (rest_tsv)")
        cur.execute("ANALYZE eval_docs")
        conn.commit()


def evaluate(conn, queries) -> dict:
    stats: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    sizes: dict[str, list[int]] = defaultdict(list)
    with conn.cursor() as cur:
        for n, q in enumerate(queries, start=1):
            gold = sorted(q["gold_ids"])
            sizes[q["level"]].append(len(gold))
            cur.execute(_RANK_SQL, {"q": grls_norm(q["query"]), "qv": q["vec"],
                                    "w": W_TSV, "k": RRF_K, "gold": gold})
            for method, rank in zip(METHODS, cur.fetchone()):
                stats[q["level"]][method].append(rank)
            if n % 100 == 0:
                print(f"    {n}/{len(queries)}", flush=True)
    return {"ranks": stats, "sizes": sizes}


def report(result: dict, index_mode: str, model: str) -> None:
    stats, sizes = result["ranks"], result["sizes"]
    width = 96
    print("\n" + "=" * width)
    print(f"индексируется: {index_mode}    модель: {model}")
    print(f"{'класс':<16}{'метод':<14}{'recall@1':>10}{'recall@5':>10}"
          f"{'recall@10':>11}{'MRR':>8}{'n':>6}{'|gold|':>8}")
    print("=" * width)
    for level in sorted(stats):
        mean_gold = sum(sizes[level]) / max(len(sizes[level]), 1)
        best = (0.0, "")
        for method in METHODS:
            ranks = stats[level][method]
            if not ranks:
                continue
            n = len(ranks)
            r1 = sum(1 for r in ranks if r <= 1) / n
            r5 = sum(1 for r in ranks if r <= 5) / n
            r10 = sum(1 for r in ranks if r <= TOP_K) / n
            mrr = sum(1.0 / r for r in ranks) / n
            best = max(best, (mrr, method))
            print(f"{level:<16}{method:<14}{r1:>10.3f}{r5:>10.3f}{r10:>11.3f}"
                  f"{mrr:>8.3f}{n:>6}{mean_gold:>8.1f}")
        if best[1]:
            print(f"{'':<16}лучший по MRR: {best[1]}")
        print("-" * width)
    print("\nrecall@k = доля запросов, где хотя бы один правильный ответ попал в топ-k.")
    print("|gold| — среднее число правильных ответов: у L6–L8 их много, поэтому")
    print("классы между собой несравнимы; сравнивать надо МЕТОДЫ внутри класса.\n")
    print("Как читать (спека engine §4.5):")
    print("  • trgm+tsv — то, что планируется в Искре; это базовая линия,")
    print("    остальные методы оцениваются тем, насколько они её обходят;")
    print("  • trgm+tsv+vec не лучше trgm+tsv → вектор не добавляет ничего поверх")
    print("    того, что и так будет: колонку embedding, HNSW и фазу эмбеддинга убрать;")
    print("  • выигрыш вектора только на L5/L6 → в режиме blob МНН и ФТГ и так внутри")
    print("    индексируемого текста, лексика их найдёт; взвесить трезво;")
    print("  • выигрыш на L2–L4 и L7 (опечатки) → вот это довод за вектор;")
    print("  • tsv проседает на L7 → ожидаемо, tsv не терпит опечаток вовсе;")
    print("    это довод держать названия под триграммами, а не под tsv;")
    print("  • hybrid_even заметно лучше hybrid → вес 0.3 из 084 под лекарства не")
    print("    подходит, подбирать свой.")


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--from-zip", metavar="FILE", help="архив ГРЛС (xlsx внутри)")
    src.add_argument("--from-db", metavar="DSN", help="Postgres medkard с grls_registry")
    p.add_argument("--index", choices=("blob", "name"), default="blob",
                   help="что индексировать: search_blob как в спеке (по умолчанию) или только название")
    p.add_argument("--limit", type=int, default=5000,
                   help="размер корпуса уникальных препаратов (0 = все; больше корпус — честнее и дороже)")
    p.add_argument("--per-level", type=int, default=150, help="запросов на класс")
    p.add_argument("--pg-dsn", default=None,
                   help="БД для временной таблицы эвала (по умолчанию POSTGRES_* из .env)")
    p.add_argument("--base-url", default=DEFAULT_BASE_URL,
                   help="по умолчанию EMBEDDING_BASE_URL / OPENAI_BASE_URL из .env")
    p.add_argument("--api-key", default=DEFAULT_API_KEY, help="по умолчанию OPENAI_API_KEY из .env")
    p.add_argument("--model", default=DEFAULT_MODEL, help="по умолчанию EMBEDDING_MODEL из .env")
    p.add_argument("--dump-queries", metavar="FILE", help="выгрузить запросы в JSON для разбора")
    args = p.parse_args()

    try:
        import psycopg
    except ImportError:
        sys.exit("нужен psycopg (v3): pip install 'psycopg[binary]'")

    rng = random.Random(SEED)
    pg_dsn = args.pg_dsn or pg_dsn_from_env()

    print(f".env: {ENV_PATH or 'не найден — переменные только из окружения'}")
    print(f"эндпоинт: {args.base_url}")
    print(f"модель:   {args.model}  (ключ: {'задан' if args.api_key else 'ПУСТ'})")
    print(f"postgres: {re.sub(r'//[^@]*@', '//***@', pg_dsn)}")
    print(f"индекс:   {args.index}")

    print("1/5 читаю корпус…", flush=True)
    rows = load_from_zip(args.from_zip) if args.from_zip else load_from_db(args.from_db)
    corpus = dedupe(rows)
    print(f"    уникальных препаратов: {len(corpus)}")
    if args.limit and len(corpus) > args.limit:
        corpus = rng.sample(corpus, args.limit)
        print(f"    корпус сокращён до {len(corpus)} (--limit)")
    docs = [search_blob(r) if args.index == "blob" else r["trade_name"] for r in corpus]
    with_ftg = sum(1 for r in corpus if len(grls_norm(r.get("pharm_group"))) >= FTG_MIN_CHARS)
    print(f"    с пригодной ФТГ: {with_ftg} ({100 * with_ftg // max(len(corpus), 1)}%)")

    print("2/5 строю запросы…", flush=True)
    queries = build_queries(corpus, args.per_level, rng)
    print(f"    запросов: {len(queries)}")
    if args.dump_queries:
        Path(args.dump_queries).write_text(json.dumps(
            [{"level": q["level"], "query": q["query"], "n_gold": len(q["gold_ids"])}
             for q in queries], ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"    выгружены: {args.dump_queries}")

    print(f"3/5 эмбеддинг корпуса и запросов моделью {args.model}…", flush=True)
    doc_vecs = embed(docs, args.base_url, args.api_key, args.model)
    q_vecs = embed([q["query"] for q in queries], args.base_url, args.api_key, args.model)
    for q, v in zip(queries, q_vecs):
        q["vec"] = _pgvector(v)
    dim = len(doc_vecs[0])
    print(f"    размерность: {dim}")

    print("4/5 временная таблица и текстовые индексы…", flush=True)
    with psycopg.connect(pg_dsn) as conn:
        build_table(conn, corpus, docs, doc_vecs, dim)
        print("5/5 ранжирую средствами Postgres…", flush=True)
        result = evaluate(conn, queries)
    # временная таблица исчезла вместе с сессией — в БД ничего не осталось
    report(result, args.index, args.model)


if __name__ == "__main__":
    main()
