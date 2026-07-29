"""
Integration tests for the pull API (api.app / api.cards) — hits the real
configured Postgres via FastAPI's TestClient. Uses one test API key scoped
to both Alenka and MDS (matching the real deployment where a key is scoped
to specific orgs), and specifies the org via ?org=.

Fixtures here are function-scoped (not module-scoped): pytest.ini only sets
asyncio_mode = auto with no asyncio_default_fixture_loop_scope, so each test
gets its own event loop — a module-scoped async fixture would be bound to
whichever loop first created it and hang when a later test's (different)
loop tries to reuse it.
"""

from __future__ import annotations

import io
import re
import sys
import uuid
from pathlib import Path

import openpyxl
import pytest
from dotenv import load_dotenv
from fastapi.testclient import TestClient

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))
load_dotenv(ROOT / ".env")

from api.app import create_app
from storage.api_keys_storage import ApiKeysStorage
from storage.organizations_storage import OrganizationsStorage

_GUID_RE = re.compile(r"GUID:\s*([0-9a-f-]{36})")


def _unique_key() -> str:
    return f"medkard_test_{uuid.uuid4().hex}"


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(create_app())


@pytest.fixture
async def alenka_org_id() -> str:
    async with OrganizationsStorage() as organizations:
        return await organizations.get_id_by_name("Alenka")


@pytest.fixture
async def mds_org_id() -> str:
    async with OrganizationsStorage() as organizations:
        return await organizations.get_id_by_name("MDS")


@pytest.fixture
async def test_key(alenka_org_id: str, mds_org_id: str) -> str:
    """A key scoped to both Alenka and MDS."""
    raw_key = _unique_key()
    async with ApiKeysStorage() as api_keys:
        key_id = await api_keys.create_key("pytest", raw_key, [alenka_org_id, mds_org_id])
    yield raw_key
    async with ApiKeysStorage() as api_keys:
        await api_keys.revoke_key(key_id)


@pytest.fixture
async def alenka_only_key(alenka_org_id: str) -> str:
    """A key scoped to Alenka only, to test out-of-scope org access."""
    raw_key = _unique_key()
    async with ApiKeysStorage() as api_keys:
        key_id = await api_keys.create_key("pytest-alenka-only", raw_key, [alenka_org_id])
    yield raw_key
    async with ApiKeysStorage() as api_keys:
        await api_keys.revoke_key(key_id)


def _auth(key: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {key}"}


def _load_workbook(content: bytes):
    return openpyxl.load_workbook(io.BytesIO(content))


def _first_guid_on(client: TestClient, key: str, org: str, date_str: str) -> str:
    resp = client.get(f"/visits/pull?date={date_str}&org={org}", headers=_auth(key))
    ws = _load_workbook(resp.content).active
    card_data = ws.cell(row=2, column=3).value
    return _GUID_RE.search(card_data).group(1)


def test_missing_key_is_rejected(client: TestClient):
    resp = client.get("/visits/check?date=2026-06-21&org=Alenka")
    assert resp.status_code == 403 or resp.status_code == 401


def test_wrong_key_is_rejected(client: TestClient):
    resp = client.get("/visits/check?date=2026-06-21&org=Alenka", headers=_auth("medkard_bogus"))
    assert resp.status_code == 401


def test_unknown_org_is_404(client: TestClient, test_key: str):
    resp = client.get("/visits/check?date=2026-06-21&org=Nonexistent", headers=_auth(test_key))
    assert resp.status_code == 404


def test_key_not_scoped_to_org_is_403(client: TestClient, alenka_only_key: str):
    resp = client.get("/visits/check?date=2026-06-21&org=MDS", headers=_auth(alenka_only_key))
    assert resp.status_code == 403


def test_check_returns_count_for_known_date(client: TestClient, test_key: str):
    resp = client.get("/visits/check?date=2026-06-21&org=Alenka", headers=_auth(test_key))
    assert resp.status_code == 200
    body = resp.json()
    assert body["date"] == "2026-06-21"
    assert isinstance(body["count"], int)
    assert body["count"] > 0


def test_org_param_is_case_insensitive(client: TestClient, test_key: str):
    exact = client.get("/visits/check?date=2026-06-21&org=Alenka", headers=_auth(test_key)).json()
    lower = client.get("/visits/check?date=2026-06-21&org=alenka", headers=_auth(test_key)).json()
    upper = client.get("/visits/check?date=2026-06-21&org=ALENKA", headers=_auth(test_key)).json()
    assert exact["count"] == lower["count"] == upper["count"]

    # filename always uses the DB's canonical casing, regardless of what the client sent
    resp = client.get("/visits/pull?date=2026-06-21&org=alenka", headers=_auth(test_key))
    assert 'filename="report_Alenka_2026-06-21.xlsx"' in resp.headers["content-disposition"]


def test_pull_returns_xlsx_file_with_one_row_per_card(client: TestClient, test_key: str):
    check_resp = client.get("/visits/check?date=2026-06-21&org=Alenka", headers=_auth(test_key))
    expected_count = check_resp.json()["count"]

    resp = client.get("/visits/pull?date=2026-06-21&org=Alenka", headers=_auth(test_key))
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]
    assert 'filename="report_Alenka_2026-06-21.xlsx"' in resp.headers["content-disposition"]

    wb = _load_workbook(resp.content)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert header[0] == "Специализация"
    assert ws.max_row - 1 == expected_count  # minus header row


def test_org_param_scopes_results(client: TestClient, test_key: str):
    alenka_count = client.get("/visits/check?date=2026-06-21&org=Alenka", headers=_auth(test_key)).json()["count"]
    mds_count = client.get("/visits/check?date=2026-06-21&org=MDS", headers=_auth(test_key)).json()["count"]

    alenka_guid = _first_guid_on(client, test_key, "Alenka", "2026-06-21")
    resp = client.get("/visits/pull?date=2026-06-21&org=MDS", headers=_auth(test_key))
    wb = _load_workbook(resp.content)
    ws = wb.active
    guids_in_mds = {ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)}
    assert not any(alenka_guid in (cell or "") for cell in guids_in_mds)  # MDS's cards don't include Alenka's guid
    assert alenka_count > 0 and mds_count >= 0


def test_check_empty_for_date_with_no_cards(client: TestClient, test_key: str):
    resp = client.get("/visits/check?date=1999-01-01&org=Alenka", headers=_auth(test_key))
    assert resp.status_code == 200
    assert resp.json()["count"] == 0


def test_pull_is_404_for_date_with_no_cards(client: TestClient, test_key: str):
    resp = client.get("/visits/pull?date=1999-01-01&org=Alenka", headers=_auth(test_key))
    assert resp.status_code == 404
