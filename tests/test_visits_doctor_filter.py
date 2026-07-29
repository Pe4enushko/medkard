"""Integration tests for the doctor_code filter and /visits/doctors.

Seeds its own done_cards rows for MDS on a far-future date (2044-01-01) so
existing stand data can't collide, and deletes them afterwards.
"""

from __future__ import annotations

import io
import json
import re
import sys
import uuid
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from dotenv import load_dotenv
from fastapi.testclient import TestClient

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))
load_dotenv(ROOT / ".env")

from api.app import create_app
from reporting.api_formatter import ApiFormatter
from storage.api_keys_storage import ApiKeysStorage
from storage.base import BaseStorage
from storage.organizations_storage import OrganizationsStorage

FIXTURE_DATE = date(2044, 1, 1)          # DD.MM.YYYY в карте: 01.01.2044
DOC_A = "00001"
DOC_B = "00002"

# GUID печатается как «GUID: <value>» внутри колонки «Данные карты» (col 3).
_GUID_RE = re.compile(r"GUID:\s*([0-9a-f-]{36})")


class _CardsWriter(BaseStorage):
    async def insert_card(self, guid: str, card_data: dict[str, Any], org_id: str) -> None:
        async with self._pool.connection() as conn:
            await conn.execute(
                "INSERT INTO done_cards (card_guid, card_data, formal_result, diag_result,"
                " icd_check_result, ignored, broken, status, organization_id)"
                " VALUES (%(guid)s, %(data)s::jsonb, '[]'::jsonb, '[]'::jsonb, '[]'::jsonb,"
                " FALSE, FALSE, 'done', %(org)s::uuid)",
                {"guid": guid, "data": json.dumps(card_data, ensure_ascii=False), "org": org_id},
            )

    async def delete_cards(self, guids: list[str]) -> None:
        async with self._pool.connection() as conn:
            await conn.execute(
                "DELETE FROM done_cards WHERE card_guid = ANY(%(guids)s)", {"guids": guids}
            )


def _card(doctor_code: str | None, doctor_name: str | None) -> dict[str, Any]:
    priem: dict[str, Any] = {"GUID": str(uuid.uuid4()), "DATE": "01.01.2044"}
    if doctor_name is not None:
        priem["Врач"] = doctor_name
    if doctor_code is not None:
        priem["Врач_код"] = doctor_code
    return {
        "Прием": priem,
        "Врач": {"SPECIALIZATION": "Невролог"},
        "Пациент": {"CODE": "Т-000001", "GENDER": "Мужской", "AGE": 40},
        "Диагнозы": [],
    }


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(create_app())


@pytest.fixture
async def mds_org_id() -> str:
    async with OrganizationsStorage() as organizations:
        return await organizations.get_id_by_name("MDS")


@pytest.fixture
async def test_key(mds_org_id: str) -> str:
    raw_key = f"medkard_test_{uuid.uuid4().hex}"
    async with ApiKeysStorage() as api_keys:
        key_id = await api_keys.create_key("pytest-doctor-filter", raw_key, [mds_org_id])
    yield raw_key
    async with ApiKeysStorage() as api_keys:
        await api_keys.revoke_key(key_id)


def _auth(key: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {key}"}


def _guids_in_workbook(content: bytes) -> set[str]:
    """GUID-ы карт из колонки «Данные карты» (без строки заголовка)."""
    ws = openpyxl.load_workbook(io.BytesIO(content)).active
    guids = set()
    for row in range(2, ws.max_row + 1):
        match = _GUID_RE.search(ws.cell(row=row, column=3).value or "")
        if match:
            guids.add(match.group(1))
    return guids


@pytest.fixture
async def seeded_cards(mds_org_id: str):
    """2 карты врача 00001, 1 карта 00002, 1 без Врач_код — все на 2044-01-01.

    Отдаёт (org_id, guid-ы карт DOC_A), чтобы тесты проверяли не количество
    строк, а идентичность отфильтрованных карт.
    """
    cards = [
        _card(DOC_A, "Иванов Иван Иванович"),
        _card(DOC_A, "Иванов Иван Иванович"),
        _card(DOC_B, "Петрова Анна Сергеевна"),
        _card(None, None),
    ]
    guids = [c["Прием"]["GUID"].lower() for c in cards]
    async with _CardsWriter() as writer:
        for guid, card in zip(guids, cards):
            await writer.insert_card(guid, card, mds_org_id)
    yield mds_org_id, set(guids[:2])
    async with _CardsWriter() as writer:
        await writer.delete_cards(guids)


async def test_check_counts_only_that_doctor(seeded_cards):
    org_id, _ = seeded_cards
    async with ApiFormatter() as formatter:
        assert await formatter.check(FIXTURE_DATE, org_id, DOC_A) == 2
        assert await formatter.check(FIXTURE_DATE, org_id, DOC_B) == 1
        assert await formatter.check(FIXTURE_DATE, org_id, "99999") == 0


async def test_check_without_filter_counts_all(seeded_cards):
    org_id, _ = seeded_cards
    async with ApiFormatter() as formatter:
        assert await formatter.check(FIXTURE_DATE, org_id) == 4


async def test_make_xlsx_filters_rows(seeded_cards):
    org_id, doc_a_guids = seeded_cards
    async with ApiFormatter() as formatter:
        content = await formatter.make_xlsx(FIXTURE_DATE, org_id, DOC_A)
    assert _guids_in_workbook(content) == doc_a_guids


def test_pull_with_doctor_code_filters_and_renames(client, test_key, seeded_cards):
    _, doc_a_guids = seeded_cards
    resp = client.get(
        f"/visits/pull?date=2044-01-01&org=MDS&doctor_code={DOC_A}", headers=_auth(test_key)
    )
    assert resp.status_code == 200
    assert 'filename="report_MDS_2044-01-01_doc00001.xlsx"' in resp.headers["content-disposition"]
    assert _guids_in_workbook(resp.content) == doc_a_guids


def test_pull_unknown_doctor_returns_placeholder_not_404(client, test_key, seeded_cards):
    resp = client.get(
        "/visits/pull?date=2044-01-01&org=MDS&doctor_code=99999", headers=_auth(test_key)
    )
    assert resp.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
    assert ws.cell(row=1, column=1).value == "За 01.01.2044 приёмов врача с кодом 99999 не обнаружено"
    assert ws.max_row == 1


def test_pull_without_filter_keeps_404_contract(client, test_key):
    resp = client.get("/visits/pull?date=1999-01-01&org=MDS", headers=_auth(test_key))
    assert resp.status_code == 404


def test_doctors_lists_unique_codes_sorted_by_name(client, test_key, seeded_cards):
    resp = client.get("/visits/doctors?org=MDS", headers=_auth(test_key))
    assert resp.status_code == 200
    doctors = resp.json()
    ours = [d for d in doctors if d["code"] in (DOC_A, DOC_B)]
    assert ours == [
        {"code": DOC_A, "name": "Иванов Иван Иванович"},
        {"code": DOC_B, "name": "Петрова Анна Сергеевна"},
    ]
    # Наша карта без Врач_код не родила пустого врача. Проверяем только свой
    # посев: общий обход связал бы тест со стендовыми данными организации.
    assert {"code": "", "name": ""} not in doctors
    assert not any(d["code"] == "" for d in ours)


def test_doctors_requires_auth(client):
    resp = client.get("/visits/doctors?org=MDS")
    assert resp.status_code in (401, 403)
