"""Build minimal GRLS-like xlsx sheets for tests (layout of the 2026-08-17 export)."""
from __future__ import annotations

from pathlib import Path

import openpyxl

HEADERS = (
    "Номер регистрационного удостоверения",
    "Дата регистрации",
    "Дата окончания действия регистрационного удостоверения",
    "Дата аннулирования регистрацион- ного удостоверения",
    "Юридическое лицо, на имя которого выдано регистрационное удостоверение",
    None,
    "Торговое наименование\nлекарственного препарата",
    "Международное непатентованное или химическое наименование",
    "Формы выпуска",
    "Сведения о стадиях производства",
    "Нормативная документация",
    "Фармако-терапевтическая группа",
    "Наличие лекарственного препарата в перечне ЖНВЛП",
    "Наличие в лекарственном препарате наркотических средств, психотропных веществ",
    "Орфанный",
)


def sample_row(**over) -> tuple:
    base = {
        "reg_number": "ЛП-000001", "registered_at": "01.02.2020", "expires_at": "",
        "annulled_at": "", "holder": 'ООО "Тест"', "holder_country": "Россия",
        "trade_name": "Тестин®", "inn_name": "тестамол",
        "forms_raw": "таблетки, 5 мг, 10 шт. - блистеры - пачки картонные - По рецепту; ",
        "production_stages": "Все стадии, ООО Тест, Россия",
        "normative_docs": "ЛП-000001-010220_x000D_\nИзм. №1",
        "pharm_group": "анальгетик", "is_vital": "Да", "narcotic": "~", "is_orphan": "",
    }
    base.update(over)
    return (base["reg_number"], base["registered_at"], base["expires_at"], base["annulled_at"],
            base["holder"], base["holder_country"], base["trade_name"], base["inn_name"],
            base["forms_raw"], base["production_stages"], base["normative_docs"],
            base["pharm_group"], base["is_vital"], base["narcotic"], base["is_orphan"])


def make_sheet(path: Path, status: str, rows: list[tuple], *,
               registry_date: str = "17.08.2026", trailer: bool = True,
               headers: tuple = HEADERS, title: str | None = None) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    # rows 1-2 empty; row 3 title in column D; row 4 empty; row 5 headers from C; row 6 marker in C
    ws.cell(row=3, column=4, value=title if title is not None
            else f"Государственный реестр лекарственных средств\nпо состоянию на {registry_date}")
    for i, h in enumerate(headers):
        if h is not None:
            ws.cell(row=5, column=3 + i, value=h)
    ws.cell(row=6, column=3, value=status)
    r = 7
    for row in rows:
        for i, v in enumerate(row):
            if v is not None:
                ws.cell(row=r, column=3 + i, value=v)
        r += 1
    if trailer:
        ws.cell(row=r, column=3, value=f"{registry_date} 05:00:00")
    wb.save(path)
    return path
