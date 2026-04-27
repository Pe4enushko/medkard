import asyncio
import sys
import types
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


@dataclass
class _FormalFinding:
    flag: str
    issue: str


@dataclass
class _FormalStructureResult:
    findings: list[_FormalFinding] = field(default_factory=list)

    @property
    def flags(self):
        return [finding.flag for finding in self.findings]

    def pretty_format(self):
        return "  Formal structure: OK"


@dataclass
class _DiagnosisResult:
    icd_code: str
    issues: list = field(default_factory=list)


@dataclass
class _Result:
    input: dict
    formal: _FormalStructureResult = field(default_factory=_FormalStructureResult)
    diagnosis: list[_DiagnosisResult] = field(default_factory=list)
    id: str | None = None


@dataclass
class _DiagnisisIssue:
    issue: str
    sources: list = field(default_factory=list)


storage_package = types.ModuleType("storage")
storage_package.__path__ = []
storage_models_package = types.ModuleType("storage.models")
storage_models_package.__path__ = []
storage_result_module = types.ModuleType("storage.models.result")
storage_result_module.DiagnosisResult = _DiagnosisResult
storage_result_module.Result = _Result
storage_result_module.FormalFinding = _FormalFinding
storage_result_module.FormalStructureResult = _FormalStructureResult
storage_result_module.DiagnisisIssue = _DiagnisisIssue
storage_models_package.DiagnosisResult = _DiagnosisResult
storage_models_package.Result = _Result
storage_models_package.FormalFinding = _FormalFinding
storage_models_package.FormalStructureResult = _FormalStructureResult
storage_models_package.DiagnisisIssue = _DiagnisisIssue
sys.modules["storage"] = storage_package
sys.modules["storage.models"] = storage_models_package
sys.modules["storage.models.result"] = storage_result_module

diagnosis_package = types.ModuleType("audit.diagnosis")
diagnosis_package.__path__ = []
diagnosis_validator_module = types.ModuleType("audit.diagnosis.validator")
diagnosis_validator_module.DiagnosisValidator = object
diagnosis_package.DiagnosisValidator = object
sys.modules["audit.diagnosis"] = diagnosis_package
sys.modules["audit.diagnosis.validator"] = diagnosis_validator_module

formal_package = types.ModuleType("audit.formal_structure")
formal_package.__path__ = []
formal_validator_module = types.ModuleType("audit.formal_structure.validator")
formal_validator_module.FormalValidator = object
formal_package.FormalValidator = object
sys.modules["audit.formal_structure"] = formal_package
sys.modules["audit.formal_structure.validator"] = formal_validator_module

import audit.pipeline as pipeline_module
from audit.models import DiagnosisAuditResult


class _FakeFormalValidator:
    async def validate(self, visit):
        return []


class _FakeDiagnosisValidator:
    def __init__(self, visit):
        self.visit = visit

    async def validate_diagnosis(self, diagnosis):
        return DiagnosisAuditResult(
            guideline_file_id=f"guideline-{diagnosis['КодМКБ']}",
            icd_code=diagnosis["КодМКБ"],
        )


def test_audit_visit_writes_one_excel_row_with_all_diagnoses(tmp_path, monkeypatch):
    monkeypatch.setattr(pipeline_module, "FormalValidator", _FakeFormalValidator)
    monkeypatch.setattr(pipeline_module, "DiagnosisValidator", _FakeDiagnosisValidator)

    visit = {
        "Прием": {"GUID": "visit-1"},
        "Пациент": {"Возраст": 10},
        "Диагнозы": [
            {"КодМКБ": "A01", "НаименованиеМКБ": "First"},
            {"КодМКБ": "B02", "НаименованиеМКБ": "Second"},
        ],
    }
    excel_path = tmp_path / "audit.xlsx"

    result = asyncio.run(pipeline_module.AuditPipeline(excel_path)._audit_visit(visit))

    assert [dx.icd_code for dx in result.diagnosis] == ["A01", "B02"]

    wb = openpyxl.load_workbook(excel_path)
    try:
        rows = list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()

    assert len(rows) == 2
    assert "icd_code=A01" in rows[1][2]
    assert "icd_code=B02" in rows[1][2]


def test_run_skips_visits_with_done_guids(tmp_path, monkeypatch):
    monkeypatch.setattr(pipeline_module, "FormalValidator", _FakeFormalValidator)
    monkeypatch.setattr(pipeline_module, "DiagnosisValidator", _FakeDiagnosisValidator)

    payload = {
        "appointments": [
            {
                "Прием": {"GUID": "0b4121b2-39e1-11f1-a224-00155daa6107"},
                "Диагнозы": [{"КодМКБ": "A01"}],
            },
            {
                "Прием": {"GUID": "2a3b5100-39e1-11f1-a224-00155daa6107"},
                "Диагнозы": [{"КодМКБ": "B02"}],
            },
        ]
    }
    excel_path = tmp_path / "audit.xlsx"

    results = asyncio.run(
        pipeline_module.AuditPipeline(excel_path).run(
            payload,
            done_guids={"0B4121B2-39E1-11F1-A224-00155DAA6107"},
        )
    )

    assert len(results) == 1
    assert results[0].input["Прием"]["GUID"] == "2a3b5100-39e1-11f1-a224-00155daa6107"

    wb = openpyxl.load_workbook(excel_path)
    try:
        rows = list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()

    assert len(rows) == 2
    assert "2a3b5100-39e1-11f1-a224-00155daa6107" in rows[1][0]


def test_run_batched_filters_done_guids_before_processing(tmp_path, monkeypatch):
    monkeypatch.setattr(pipeline_module, "FormalValidator", _FakeFormalValidator)
    monkeypatch.setattr(pipeline_module, "DiagnosisValidator", _FakeDiagnosisValidator)

    payload = {
        "appointments": [
            {
                "Прием": {"GUID": "11111111-39e1-11f1-a224-00155daa6107"},
                "Диагнозы": [{"КодМКБ": "A01"}],
            },
            {
                "Прием": {"GUID": "22222222-39e1-11f1-a224-00155daa6107"},
                "Диагнозы": [{"КодМКБ": "B02"}],
            },
            {
                "Прием": {"GUID": "33333333-39e1-11f1-a224-00155daa6107"},
                "Диагнозы": [{"КодМКБ": "C03"}],
            },
        ]
    }
    excel_path = tmp_path / "audit.xlsx"

    results = asyncio.run(
        pipeline_module.AuditPipeline(excel_path).run_batched(
            payload,
            num_batches=2,
            done_guids={"22222222-39e1-11f1-a224-00155daa6107"},
        )
    )

    assert [result.input["Прием"]["GUID"] for result in results] == [
        "11111111-39e1-11f1-a224-00155daa6107",
        "33333333-39e1-11f1-a224-00155daa6107",
    ]

    wb = openpyxl.load_workbook(excel_path)
    try:
        rows = list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()

    assert len(rows) == 3
    assert "11111111-39e1-11f1-a224-00155daa6107" in rows[1][0]
    assert "33333333-39e1-11f1-a224-00155daa6107" in rows[2][0]
