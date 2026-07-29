"""build_empty_report_bytes: single-cell placeholder workbook."""
import io

import openpyxl

from parsers.excel import build_empty_report_bytes


def test_empty_report_has_message_in_a1_and_nothing_else():
    message = "За 01.01.2044 приёмов врача с кодом 00001 не обнаружено"
    content = build_empty_report_bytes(message)
    ws = openpyxl.load_workbook(io.BytesIO(content)).active
    assert ws.cell(row=1, column=1).value == message
    assert ws.max_row == 1 and ws.max_column == 1
