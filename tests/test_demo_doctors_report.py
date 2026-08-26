"""The made-up doctor must not reach the xlsx report that goes to FTP.

The report is what another product shows on the demo, and there it has to
look exactly as it did before the crutch: no doctor at all. See
src/api/demo_doctors.py for the stamp itself.
"""

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from api import demo_doctors  # noqa: E402
from audit.excel_formatter import ExcelFormatter  # noqa: E402

STAMPED = {
    "Пациент": {"Код": "к0138172"},
    "Врач": {"SPECIALIZATION": "Педиатр"},
    "Прием": {
        "NUM": "ЦДЗ-00857211",
        "DATE": "24.08.2026",
        "GUID": "e92120c0-9682-42ac-9188-d92afbefe982",
        "Врач": "Врач 90001",
        "Врач_код": "90001",
    },
}


def _card(priem: dict | None = None) -> dict:
    card = {"Пациент": {"Код": "к0138172"}, "Врач": {"SPECIALIZATION": "Педиатр"}}
    if priem is not None:
        card["Прием"] = priem
    return card


# ── the pure function ────────────────────────────────────────────────────────

def test_unstamp_drops_the_doctor_and_keeps_the_rest():
    priem = demo_doctors.unstamp(STAMPED)["Прием"]
    assert "Врач" not in priem
    assert "Врач_код" not in priem
    assert priem["NUM"] == "ЦДЗ-00857211"
    assert priem["DATE"] == "24.08.2026"
    assert priem["GUID"] == "e92120c0-9682-42ac-9188-d92afbefe982"


def test_unstamp_keeps_the_specialization_block():
    # Специализация in the report comes from the top-level Врач dict 1C sends;
    # the crutch never touched it and neither may the strip.
    assert demo_doctors.unstamp(STAMPED)["Врач"] == {"SPECIALIZATION": "Педиатр"}


def test_unstamp_does_not_mutate_the_card_it_was_given():
    card = _card({"GUID": "g1", "Врач": "Врач 90001", "Врач_код": "90001"})
    demo_doctors.unstamp(card)
    assert card["Прием"]["Врач_код"] == "90001"


def test_unstamp_leaves_a_card_without_a_doctor_alone():
    card = _card({"GUID": "g1", "DATE": "24.08.2026"})
    assert demo_doctors.unstamp(card) == card


def test_unstamp_survives_a_card_without_a_priem_block():
    card = _card()
    assert demo_doctors.unstamp(card) == card


# ── the wiring: ExcelFormatter → the report file ─────────────────────────────

def _report_card_text(tmp_path: Path, org_name: str) -> tuple[str, str]:
    """Write one stamped card through the formatter, return (специализация, данные карты)."""
    path = tmp_path / "report.xlsx"
    formatter = ExcelFormatter(path, org_name=org_name)
    formatter._write_rows(
        [{
            "id": 1,
            "card_guid": STAMPED["Прием"]["GUID"],
            "card_data": STAMPED,
            "formal_result": None,
            "diag_result": None,
            "icd_check_result": None,
        }],
        {},
    )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        return str(rows[0][0]), str(rows[0][2])
    finally:
        wb.close()


def test_report_of_the_flagged_org_carries_no_doctor(tmp_path, monkeypatch):
    monkeypatch.setenv("DEMO_DOCTOR_STAMP_ORG", "Alenka")
    specialization, card_text = _report_card_text(tmp_path, "Alenka")
    assert "Врач_код" not in card_text
    assert "Врач 90001" not in card_text
    assert specialization == "Педиатр"


def test_report_of_another_org_keeps_the_doctor(tmp_path, monkeypatch):
    monkeypatch.setenv("DEMO_DOCTOR_STAMP_ORG", "Alenka")
    _, card_text = _report_card_text(tmp_path, "MDS")
    assert "Врач_код: 90001" in card_text


def test_report_keeps_the_doctor_when_the_crutch_is_off(tmp_path, monkeypatch):
    monkeypatch.delenv("DEMO_DOCTOR_STAMP_ORG", raising=False)
    _, card_text = _report_card_text(tmp_path, "Alenka")
    assert "Врач_код: 90001" in card_text


def test_formatter_without_an_org_name_keeps_the_doctor(tmp_path, monkeypatch):
    # export_guids / export_all have no org to check against; they must not
    # silently strip anything.
    monkeypatch.setenv("DEMO_DOCTOR_STAMP_ORG", "Alenka")
    path = tmp_path / "report.xlsx"
    formatter = ExcelFormatter(path)
    formatter._write_rows(
        [{
            "id": 1,
            "card_guid": "g1",
            "card_data": STAMPED,
            "formal_result": None,
            "diag_result": None,
            "icd_check_result": None,
        }],
        {},
    )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        card_text = str(list(wb.active.iter_rows(min_row=2, values_only=True))[0][2])
    finally:
        wb.close()
    assert "Врач_код: 90001" in card_text


# ── the wiring: report scripts → ExcelFormatter ──────────────────────────────

import ast  # noqa: E402


@pytest.mark.parametrize("script", [
    "scripts/audit-one-c-period.py",
    "scripts/operator/send_report_ftp.py",
    "scripts/operator/create_report.py",
])
def test_report_scripts_tell_the_formatter_which_org_they_export(script):
    # These three build the xlsx another product reads. A formatter built
    # without org_name strips nothing, so the crutch would leak silently.
    tree = ast.parse((ROOT / script).read_text(encoding="utf-8"))
    calls = [
        node for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and getattr(node.func, "id", None) == "ExcelFormatter"
    ]
    assert calls, f"{script}: no ExcelFormatter call found"
    for call in calls:
        assert any(kw.arg == "org_name" for kw in call.keywords), \
            f"{script}: ExcelFormatter built without org_name"
