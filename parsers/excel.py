"""
excel.py — append audit results to an xlsx workbook.

Each row contains three columns with human-readable text:
  - ``input``            — raw visit payload (source JSON from 1C)
  - ``formal_structure`` — FormalStructureResult
  - ``diagnosis``        — DiagnosisAuditResult or a list of them

Usage::
    from parsers.excel import AuditExcelWriter
    from audit.models import DiagnosisAuditResult, FormalStructureResult

    writer = AuditExcelWriter("results.xlsx")
    writer.append(
        visit=raw_visit_dict,
        formal=FormalStructureResult(...),
        diagnosis=DiagnosisAuditResult(...),
    )
"""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from audit.models import DiagnosisAuditResult, FormalStructureResult

_HEADERS = ["input", "formal_structure", "diagnosis"]
logger = logging.getLogger(__name__)


def _pretty(obj: Any) -> str:
    if hasattr(obj, "pretty_format") and callable(obj.pretty_format):
        return obj.pretty_format()

    if isinstance(obj, list) and all(
        hasattr(item, "pretty_format") and callable(item.pretty_format)
        for item in obj
    ):
        if not obj:
            return "—"
        return "\n\n".join(
            f"{idx}.\n{item.pretty_format()}"
            for idx, item in enumerate(obj, start=1)
        )

    if is_dataclass(obj):
        obj = asdict(obj)

    return _format_value(obj)


def _format_value(value: Any, indent: int = 0) -> str:
    prefix = " " * indent

    if is_dataclass(value):
        value = asdict(value)

    if isinstance(value, dict):
        if not value:
            return f"{prefix}—"

        lines: list[str] = []
        for key, item in value.items():
            label = str(key)
            if _is_scalar(item):
                lines.append(f"{prefix}{label}: {_format_scalar(item)}")
            else:
                lines.append(f"{prefix}{label}:")
                lines.append(_format_value(item, indent + 2))
        return "\n".join(lines)

    if isinstance(value, list):
        if not value:
            return f"{prefix}—"

        lines: list[str] = []
        for idx, item in enumerate(value, start=1):
            if _is_scalar(item):
                lines.append(f"{prefix}{idx}. {_format_scalar(item)}")
            else:
                lines.append(f"{prefix}{idx}.")
                lines.append(_format_value(item, indent + 2))
        return "\n".join(lines)

    return f"{prefix}{_format_scalar(value)}"


def _is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _format_scalar(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "да" if value else "нет"
    return str(value)


class AuditExcelWriter:
    """Append audit results to an xlsx file, creating it with a header row if absent.

    Args:
        path: Path to the xlsx output file (created automatically if missing).
    """

    def __init__(self, path: str | Path) -> None:
        self._path = Path(path)

    def _open_or_create(self) -> tuple[Workbook, Worksheet]:
        if self._path.exists():
            wb = openpyxl.load_workbook(self._path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
        if ws.cell(row=1, column=4).value == "sources":
            ws.delete_cols(4, 1)
        for idx, header in enumerate(_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
        return wb, ws  # type: ignore[return-value]

    def append(
        self,
        visit: dict[str, Any],
        formal: FormalStructureResult,
        diagnosis: DiagnosisAuditResult | list[DiagnosisAuditResult],
    ) -> None:
        """Append one result row and save the workbook.

        Args:
            visit:     Raw visit dict (source JSON from 1C).
            formal:    Formal-structure audit result.
            diagnosis: Diagnosis audit result(s) for this visit.
        """
        try:
            row = [
                _pretty(visit),
                _pretty(formal),
                _pretty(diagnosis),
            ]
            wb, ws = self._open_or_create()
            ws.append(row)
            wb.save(self._path)
            logger.info("📊 EXCEL APPEND OK path=%s", self._path)
        except Exception:
            logger.exception("📊 EXCEL APPEND FAILED path=%s", self._path)
            raise

    def rows_count(self) -> int:
        """Return current number of rows in the active worksheet (0 if file absent)."""
        if not self._path.exists():
            return 0
        wb = openpyxl.load_workbook(self._path)
        try:
            return wb.active.max_row
        finally:
            wb.close()
