"""Чтение реестра ГРЛС и производные тексты записи.

Нормализация и производные формы — зеркало канона medkard (спека §3.1 и
src/grls/normalize.py). Держать идентичными: расхождение здесь означает, что
эвал меряет не то, что будет в проде.
"""
from __future__ import annotations

import re
import sys
import unicodedata
import zipfile
from pathlib import Path

# ФТГ короче — обрывки вроде «~» или «прочие»; в классы по ФТГ не берём.
FTG_MIN_CHARS = 12
# Порог для класса «МНН → торговое»: пары с триграммным сходством выше —
# лексические («Церебролизин» ← «церебролизин»), семантику не проверяют.
INN_LEXICAL_MAX = 0.30

_JUNK = re.compile(r"[®™©\"«»„“”'`]")
_SPACES = re.compile(r"\s+")
_WORD_SPLIT = re.compile(r"[^0-9a-zа-я]+")


def grls_norm(text: str | None) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKC", text)
    text = _JUNK.sub("", text)
    text = text.replace("ё", "е").replace("Ё", "Е").replace("~", "")
    return _SPACES.sub(" ", text).strip().lower()


def _trigrams(text: str) -> set[str]:
    out: set[str] = set()
    for word in _WORD_SPLIT.split(grls_norm(text)):
        if not word:
            continue
        padded = f"  {word} "
        for i in range(len(padded) - 2):
            out.add(padded[i:i + 3])
    return out


def trgm_similarity(a: str, b_trg: set[str]) -> float:
    """Приближение similarity() — только для ОТБОРА пар при сборке запросов.

    Ранжирование считает Postgres; это приближение в нём не участвует.
    """
    a_trg = _trigrams(a)
    if not a_trg or not b_trg:
        return 0.0
    inter = len(a_trg & b_trg)
    return inter / len(a_trg | b_trg) if inter else 0.0


# ─────────────────────────── производные формы ───────────────────────────
# Первый сегмент до запятой = лекарственная форма, последний после ' - ' =
# условия отпуска.
def split_forms(forms_raw: str | None) -> list[str]:
    return [p.strip() for p in (forms_raw or "").split(";") if p.strip()]


def _unique(items: list[str]) -> list[str]:
    seen, out = set(), []
    for it in items:
        if it and it not in seen:
            seen.add(it)
            out.append(it)
    return out


def derive_dosage_forms(forms: list[str]) -> list[str]:
    return _unique([el.split(",", 1)[0].strip() for el in forms if not el.startswith("-")])


def derive_dispensing(forms: list[str]) -> list[str]:
    return _unique([el.rsplit(" - ", 1)[1].strip() for el in forms if " - " in el])


def names_text(rec: dict) -> str:
    """Названия: торговое + МНН. По ним в проде идут триграммы."""
    return " | ".join(p.strip() for p in (rec.get("trade_name"), rec.get("inn_name"))
                      if p and p.strip() and p.strip() != "~")


def rest_text(rec: dict) -> str:
    """Всё остальное: ФТГ, лекформы, отпуск, держатель. По нему в проде идёт tsv."""
    forms = split_forms(rec.get("forms_raw"))
    parts = [
        rec.get("pharm_group"),
        ", ".join(derive_dosage_forms(forms)[:6]),
        ", ".join(derive_dispensing(forms)[:4]),
        rec.get("holder"),
    ]
    return " | ".join(p.strip() for p in parts if p and p.strip() and p.strip() != "~")


def search_blob(rec: dict) -> str:
    """Поисковый текст записи целиком — как search_blob() в спеке engine §4.5.

    forms_raw целиком НЕ идёт: у препарата с десятком упаковок это килобайты,
    размывающие и лексический, и векторный сигнал; суть уже в производных.
    """
    return " | ".join(p for p in (names_text(rec), rest_text(rec)) if p)


# ─────────────────────────── чтение корпуса ───────────────────────────
# Раскладка листа — как в парсере medkard: строка 5 заголовки, данные с 7-й,
# колонки C..Q. Внутри среза: 4 держатель, 6 торговое, 7 МНН, 8 формы, 11 ФТГ.
_COL_HOLDER, _COL_TRADE, _COL_INN, _COL_FORMS, _COL_FTG = 4, 6, 7, 8, 11


def load_from_zip(path: str, quiet: bool = False) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("нужен openpyxl: pip install openpyxl")

    first, n_cols = 2, 15
    rows: list[dict] = []
    with zipfile.ZipFile(path) as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not names:
            sys.exit(f"в архиве {path} нет xlsx")
        tmp = Path("/tmp/grls_eval_sheets")
        tmp.mkdir(exist_ok=True)
        for name in names:
            target = tmp / Path(name).name
            with zf.open(name) as src, open(target, "wb") as dst:
                dst.write(src.read())
            wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i < 7:
                    continue
                cells = (tuple(row or ()) + (None,) * (first + n_cols))[first:first + n_cols]

                def cell(idx: int) -> str | None:
                    v = cells[idx]
                    if v is None:
                        return None
                    v = str(v).strip()
                    return v if v and v != "~" else None

                trade = cell(_COL_TRADE)
                if not trade:
                    continue
                rows.append({
                    "trade_name": trade,
                    "inn_name": cell(_COL_INN),
                    "pharm_group": cell(_COL_FTG),
                    "forms_raw": cell(_COL_FORMS),
                    "holder": cell(_COL_HOLDER),
                })
            wb.close()
            if not quiet:
                print(f"  прочитан {Path(name).name}: всего строк {len(rows)}", flush=True)
    return rows


def load_from_db(dsn: str) -> list[dict]:
    import psycopg
    with psycopg.connect(dsn) as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT trade_name, inn_name, pharm_group, forms_raw, holder "
            "FROM grls_registry WHERE NOT is_substance AND trade_name IS NOT NULL"
        )
        cols = [d.name for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def dedupe(rows: list[dict]) -> list[dict]:
    """Один препарат = одно уникальное нормализованное торговое наименование.

    В реестре у одного названия десятки строк (разные упаковки и редакции РУ);
    для эвала поиска они шум: цель — найти ПРЕПАРАТ, а не строку. Побочный
    эффект: омонимы разных производителей схлопываются в одну запись.
    """
    seen: dict[str, dict] = {}
    for r in rows:
        key = grls_norm(r["trade_name"])
        if not key:
            continue
        cur = seen.get(key)
        if cur is None:
            seen[key] = r
        elif (bool(r.get("inn_name")) + bool(r.get("pharm_group")) >
              bool(cur.get("inn_name")) + bool(cur.get("pharm_group"))):
            seen[key] = r  # предпочитаем более полную запись
    return list(seen.values())
