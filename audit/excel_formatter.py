"""
audit/excel_formatter.py — export done_cards rows to an xlsx workbook.

Reads every row from the done_cards table (or a specific subset by guid)
and appends them to an Excel file via AuditExcelWriter.  The pipeline no
longer writes Excel directly; call this after the pipeline completes (or
any time) to regenerate / update the workbook from the DB.

Usage::
    from audit.excel_formatter import ExcelFormatter

    async with ExcelFormatter("audit_results.xlsx") as fmt:
        written = await fmt.export_all()
        print(f"wrote {written} rows")
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from parsers.excel import AuditExcelWriter
from storage.base import BaseStorage
from storage.models.result import DiagnosisResult, FormalFinding, FormalStructureResult

logger = logging.getLogger(__name__)


def _parse_formal(text: str) -> FormalStructureResult:
    """Reconstruct a minimal FormalStructureResult from its pretty_format text."""
    findings: list[FormalFinding] = []
    for line in text.splitlines():
        line = line.strip()
        if line.startswith("[") and "]" in line:
            bracket_end = line.index("]")
            flag = line[1:bracket_end]
            issue = line[bracket_end + 1:].strip().lstrip(": ")
            findings.append(FormalFinding(flag=flag, issue=issue))
    return FormalStructureResult(findings=findings)


def _parse_diagnosis(text: str) -> list[DiagnosisResult]:
    """Reconstruct DiagnosisResult list from pretty_format text (best-effort)."""
    results: list[DiagnosisResult] = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("[") and "]" in stripped and not stripped.startswith("[•"):
            bracket_end = stripped.index("]")
            code = stripped[1:bracket_end]
            if code and code != "•":
                results.append(DiagnosisResult(icd_code=code))
    return results


class _DoneCardsReader(BaseStorage):
    async def fetch_all(self) -> list[dict[str, Any]]:
        async with self._pool.connection() as conn:
            cur = await conn.execute(
                "SELECT id, card_guid, card_data, formal_result, diag_result "
                "FROM done_cards ORDER BY id"
            )
            return await cur.fetchall()

    async def fetch_by_guids(self, guids: set[str]) -> list[dict[str, Any]]:
        async with self._pool.connection() as conn:
            cur = await conn.execute(
                "SELECT id, card_guid, card_data, formal_result, diag_result "
                "FROM done_cards WHERE card_guid = ANY(%(guids)s) ORDER BY id",
                {"guids": list(guids)},
            )
            return await cur.fetchall()


def _existing_guids_in_excel(excel: AuditExcelWriter) -> set[str]:
    """Return the set of appointment GUIDs already present in the Excel sheet.

    Each input cell contains the pretty-formatted visit dict; the GUID appears
    as the value of the «GUID» key somewhere in that text.
    """
    import openpyxl

    path = excel._path
    if not path.exists():
        return set()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        existing: set[str] = set()
        input_col = None
        for cell in ws[1]:
            if cell.value == "input":
                input_col = cell.column
                break
        if input_col is None:
            return set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_value = row[input_col - 1]
            if not cell_value:
                continue
            for line in str(cell_value).splitlines():
                stripped = line.strip()
                if stripped.startswith("GUID:"):
                    guid = stripped[len("GUID:"):].strip().lower()
                    if guid:
                        existing.add(guid)
                    break
        return existing
    finally:
        wb.close()


class ExcelFormatter:
    """Async context-manager that exports done_cards rows to an xlsx file.

    Args:
        excel_path: Path to the output xlsx file (created if absent).
    """

    def __init__(self, excel_path: str | Path) -> None:
        self._excel = AuditExcelWriter(excel_path)
        self._reader = _DoneCardsReader()

    async def __aenter__(self) -> "ExcelFormatter":
        await self._reader.__aenter__()
        return self

    async def __aexit__(self, *args: object) -> None:
        await self._reader.__aexit__(*args)

    async def export_all(self) -> int:
        """Write every done_cards row to Excel. Returns number of rows written."""
        rows = await self._reader.fetch_all()
        return self._write_rows(rows)

    async def export_guids(self, guids: set[str]) -> int:
        """Write only the rows matching *guids*. Returns number of rows written."""
        rows = await self._reader.fetch_by_guids(guids)
        return self._write_rows(rows)

    def _write_rows(self, rows: list[dict[str, Any]]) -> int:
        existing = _existing_guids_in_excel(self._excel)
        written = 0
        for row in rows:
            guid = (row["card_guid"] or "").lower()
            if guid and guid in existing:
                logger.debug("📊 skipping already exported card guid=%s", guid)
                continue
            visit = json.loads(row["card_data"])
            formal = _parse_formal(row["formal_result"])
            diagnosis = _parse_diagnosis(row["diag_result"])
            self._excel.append(visit=visit, formal=formal, diagnosis=diagnosis)
            logger.debug("📊 exported card id=%s guid=%s", row["id"], guid)
            written += 1
        logger.info("📊 ExcelFormatter exported %d row(s)", written)
        return written
